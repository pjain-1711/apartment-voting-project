from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from app.models import Voter, Vote, Wing
from datetime import datetime


def create_anonymous_export():
    """Create anonymous Excel export - no voter identification"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous Votes"

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Headers
    headers = ['Counter #', 'Wing', 'Male Vote', 'Female Vote', 'Timestamp']
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    voters = Voter.query.order_by(Voter.counter_number).all()
    for row_num, voter in enumerate(voters, 2):
        votes = Vote.query.filter_by(voter_id=voter.id).all()

        male_vote = "N/A"
        female_vote = "N/A"

        for vote in votes:
            if vote.nominee.gender == 'male':
                male_vote = vote.nominee.name
            elif vote.nominee.gender == 'female':
                female_vote = vote.nominee.name

        sheet.cell(row=row_num, column=1).value = voter.counter_number
        sheet.cell(row=row_num, column=2).value = voter.wing.name
        sheet.cell(row=row_num, column=3).value = male_vote
        sheet.cell(row=row_num, column=4).value = female_vote
        sheet.cell(row=row_num, column=5).value = voter.voted_at.strftime('%Y-%m-%d %H:%M:%S')

    # Adjust column widths
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 20

    return workbook


def create_detailed_export():
    """Create detailed Excel export - includes voter identification"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Detailed Votes"

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Headers
    headers = ['Counter #', 'Flat #', 'Wing', 'Voter Name', 'Phone', 'Male Vote', 'Female Vote', 'Timestamp']
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows
    voters = Voter.query.order_by(Voter.counter_number).all()
    for row_num, voter in enumerate(voters, 2):
        votes = Vote.query.filter_by(voter_id=voter.id).all()

        male_vote = "N/A"
        female_vote = "N/A"

        for vote in votes:
            if vote.nominee.gender == 'male':
                male_vote = vote.nominee.name
            elif vote.nominee.gender == 'female':
                female_vote = vote.nominee.name

        sheet.cell(row=row_num, column=1).value = voter.counter_number
        sheet.cell(row=row_num, column=2).value = voter.flat_number
        sheet.cell(row=row_num, column=3).value = voter.wing.name
        sheet.cell(row=row_num, column=4).value = voter.name
        sheet.cell(row=row_num, column=5).value = voter.phone_number
        sheet.cell(row=row_num, column=6).value = male_vote
        sheet.cell(row=row_num, column=7).value = female_vote
        sheet.cell(row=row_num, column=8).value = voter.voted_at.strftime('%Y-%m-%d %H:%M:%S')

    # Adjust column widths
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 15
    sheet.column_dimensions['F'].width = 25
    sheet.column_dimensions['G'].width = 25
    sheet.column_dimensions['H'].width = 20

    return workbook


def create_results_export():
    """Create results Excel export"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Election Results"

    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    winner_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    # Headers
    headers = ['Wing', 'Gender', 'Nominee Name', 'Flat #', 'Vote Count', 'Rank', 'Status']
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Get all wings and their results
    from app.models import Result
    results = Result.query.join(Wing).order_by(Wing.name, Result.gender, Result.rank).all()

    for row_num, result in enumerate(results, 2):
        sheet.cell(row=row_num, column=1).value = result.wing.name
        sheet.cell(row=row_num, column=2).value = result.gender.capitalize()
        sheet.cell(row=row_num, column=3).value = result.nominee.name
        sheet.cell(row=row_num, column=4).value = result.nominee.flat_number
        sheet.cell(row=row_num, column=5).value = result.vote_count
        sheet.cell(row=row_num, column=6).value = result.rank

        status_cell = sheet.cell(row=row_num, column=7)
        status_cell.value = "WINNER" if result.is_winner else "Candidate"

        if result.is_winner:
            for col in range(1, 8):
                sheet.cell(row=row_num, column=col).fill = winner_fill

    # Adjust column widths
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 12
    sheet.column_dimensions['F'].width = 8
    sheet.column_dimensions['G'].width = 12

    return workbook
